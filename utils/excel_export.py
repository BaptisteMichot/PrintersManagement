# ==========================================
# Export Excel
# Utilitaires pour exporter les commandes en Excel basé sur le modèle
# ==========================================

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
import os


def safe_write_cell(ws, cell_ref, value):
    """
    Écrit une valeur dans une cellule en gérant les cellules fusionnées.
    Utilise la cellule principale de la région fusionnée.
    
    Args:
        ws: Feuille de travail openpyxl
        cell_ref: Référence de la cellule (ex: 'C12')
        value: Valeur à écrire
    """
    try:
        # Essayer d'écrire directement
        ws[cell_ref].value = value
    except Exception:
        try:
            # Si c'est une MergedCell, chercher la cellule principale de la fusion
            for merged_range in ws.merged_cells.ranges:
                if cell_ref in merged_range:
                    # Écrire dans la première cellule de la plage fusionnée
                    top_left = merged_range.start_cell
                    top_left.value = value
                    return
            # Si pas fusionnée, réécrire directement
            ws[cell_ref].value = value
        except Exception:
            pass  # Ignorer silencieusement les erreurs d'écriture


def write_cell_data(ws, cell_ref, value):
    """
    Écrit dans une cellule de données (lignes 23-32) sans unmerger.
    Ignore les erreurs si la cellule est fusionnée.
    
    Args:
        ws: Feuille de travail openpyxl
        cell_ref: Référence de la cellule
        value: Valeur à écrire
    """
    try:
        ws[cell_ref].value = value
    except Exception:
        pass  # Ignorer les erreurs pour les cellules fusionnées


def export_order_to_excel(order_data, filepath, originator_name="Ibrahima DIARRA"):
    """
    Exporte une commande en fichier Excel basé exactement sur le modèle ABB.
    Remplir seulement le numéro de commande, la date et les lignes de commande.
    
    Args:
        order_data (dict): Dictionnaire contenant:
            - po_number: Numéro de commande
            - date: Date de la commande
            - lines: Liste de dictionnaires avec cartridge_type, description, quantity, unit_price, total
        filepath (str): Chemin où sauvegarder le fichier Excel
        originator_name (str): Nom de la personne qui passe la commande (défaut: "Ibrahima DIARRA")
        
    Returns:
        bool: True si l'export a réussi, False sinon
    """
    try:
        # Déterminer quel modèle utiliser selon l'originator
        if originator_name == "Ibrahima DIARRA":
            template_name = 'order_model.xlsx'
        else:
            template_name = 'order_model_other_user.xlsx'
        
        # Charger le modèle approprié
        template_path = os.path.join(os.path.dirname(__file__), '..', 'assets', template_name)
        wb = load_workbook(template_path)
        ws = wb.active
        
        # Déverrouiller la feuille au cas où elle serait protégée
        ws.protection.sheet = False
        
        # Remplir seulement les informations variables
        safe_write_cell(ws, 'C12', order_data.get('po_number', 'N/A'))
        
        # Date
        safe_write_cell(ws, 'C15', order_data.get('date', datetime.now().strftime('%d/%m/%Y')))
        
        # Remplir le champ Originator (ligne 37, colonne C) SEULEMENT si ce n'est pas Ibrahima
        # Si c'est Ibrahima, le modèle contient déjà son nom
        if originator_name != "Ibrahima DIARRA":
            safe_write_cell(ws, 'C37', originator_name)
        
        # Effacer les anciennes lignes de données (lignes 23-32, pas la ligne 33 qui est le Total)
        for row_idx in range(23, 33):
            write_cell_data(ws, f'A{row_idx}', None)
            write_cell_data(ws, f'B{row_idx}', None)
            write_cell_data(ws, f'H{row_idx}', None)
            write_cell_data(ws, f'I{row_idx}', None)
        
        # Ajouter les lignes de commande
        current_row = 23
        total_amount = 0
        
        for line in order_data.get('lines', []):
            quantity = float(line.get('quantity', 0))
            unit_price = float(line.get('unit_price', 0))
            total = float(line.get('total', 0))
            
            # Sauter les lignes vides
            if quantity == 0 and not line.get('description'):
                continue
            
            # Construire la description avec cartridge_type
            description = f"{line.get('cartridge_type', '')} {line.get('description', '')}".strip()
            
            write_cell_data(ws, f'A{current_row}', int(quantity) if quantity == int(quantity) else quantity)
            write_cell_data(ws, f'B{current_row}', description)
            write_cell_data(ws, f'H{current_row}', unit_price)
            
            # Ajouter la formule de total
            write_cell_data(ws, f'I{current_row}', f'=IF(H{current_row}="","",H{current_row}*A{current_row})')
            
            # Préserver l'alignement vertical du modèle tout en appliquant wrap_text
            try:
                cell_ref = f'B{current_row}'
                cell_b = ws[cell_ref]
                existing_alignment = cell_b.alignment
                cell_b.alignment = Alignment(
                    wrap_text=True, 
                    vertical=existing_alignment.vertical if existing_alignment else 'center',
                    horizontal=existing_alignment.horizontal if existing_alignment else 'left'
                )
            except Exception:
                # Si la modification d'alignement échoue, ignorer
                pass
            
            ws.row_dimensions[current_row].height = 30
            
            total_amount += total
            current_row += 1
        
        # Remplir la ligne Total (ligne 33) avec la formule SUM
        try:
            safe_write_cell(ws, 'I33', f'=SUM(I23:I32)')
        except Exception as e:
            print(f"Cannot write to total cell: {str(e)}")
        
        # Sauvegarder
        wb.save(filepath)
        return True
        
    except Exception as e:
        print(f"Error exporting order to Excel: {str(e)}")
        raise
